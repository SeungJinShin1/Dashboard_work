import io
import csv
from openpyxl import load_workbook
from fastapi import UploadFile
from backend.services.gemini import gemini_service
import json

class ExcelProcessor:
    async def process_schedule(self, file: UploadFile) -> list:
        try:
            # 1. Read Excel file
            print("DEBUG: Reading Excel file...")
            contents = await file.read()
            wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
            ws = wb.active
            
            # Convert to CSV string for token efficiency
            output = io.StringIO()
            writer = csv.writer(output)
            for row in ws.iter_rows(values_only=True):
                # Filter out completely empty rows
                if any(cell is not None for cell in row):
                    writer.writerow(row)
            
            csv_data = output.getvalue()
            # print(f"DEBUG: Excel read successfully. CSV length: {len(csv_data)}")
            
            # 2. Prompt Gemini
            print("DEBUG: Prompting Gemini...")
            prompt = f"""
            Analyze the following school schedule data and extract events.
            
            [Data]
            {csv_data}
            
            [System Prompt]
            Extract school events from this data. Return a valid JSON list of objects with these keys:
            - title: Name of the event
            - date: YYYY-MM-DD format
            - time: Specific time (e.g. "14:00~16:00") or "All Day"
            - location: Place of the event
            - participants: Target audience (e.g. "1-6 Graders", "Teachers")
            - manager: Person in charge
            - type: "official", "trip", "personal"
            - note: Any other details (e.g. "Business Trip to Seoul")

            Rules:
            - Ignore empty rows.
            - Normalize dates to YYYY-MM-DD.
            - If a field is missing, use empty string "".
            - Return ONLY raw JSON.
            """
            
            response_text = await gemini_service.generate_content(prompt)
            print(f"DEBUG: Gemini raw response: {response_text[:100]}...") # Print first 100 chars
            
            # Check for service-level errors
            if response_text.startswith("Error:"):
                 raise Exception(f"AI Service Error: {response_text}")

            # 3. Parse JSON with Regex
            import re
            clean_text = response_text.strip()
            match = re.search(r'\[.*\]', clean_text, re.DOTALL)
            if match:
                clean_text = match.group(0)
            else:
                clean_text = clean_text.replace("```json", "").replace("```", "").strip()

            try:
                events = json.loads(clean_text)
                print(f"DEBUG: Parsed {len(events)} events")
                return events
            except json.JSONDecodeError as e:
                print(f"JSON Parse Failed. Raw Text: {response_text}")
                # Return raw text in error for debugging
                raise Exception(f"Failed to parse AI response. Raw: {response_text[:500]}... Error: {e}")

        except Exception as e:
            print(f"DEBUG: ExcelProcessor Error: {e}")
            # Propagate specific error message up to the router
            raise Exception(f"{str(e)}")

excel_processor = ExcelProcessor()
